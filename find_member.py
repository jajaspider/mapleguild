import time
import re

from openpyxl import Workbook
from selenium import webdriver

def get_member(guild_name, servername):
    # 멤버를 담을 공백 리스트 생성
    member_list = []

    url_server = "elysium"
    if servername == "엘리시움":
        url_server = "elysium"

    url = "https://maple.gg/guild/" + url_server + "/" + guild_name
    driver.get(url)
    time.sleep(3)

    try:
        # 정보갱신버튼을 누르도록 함
        driver.find_element_by_xpath('//*[@id="btn-sync"]').click()
        time.sleep(1)

        # 길드 갱신 확인 메세지박스의 확인을 클릭합니다.
        comfirm_alert = driver.find_element_by_id('simpleAlert')
        comfirm_alert.click()
        time.sleep(5)
    except Exception as e:
        print(e)
        time.sleep(5)

    try:
        # maple.gg의 guild-content 블록을 가져오도록 함
        guild_data = driver.find_element_by_id('guild-content')
        # a 태그에는 길드원의 닉네임과 공백이 담겨있음
        guild_member_name = guild_data.find_elements_by_tag_name('a')

        for i in guild_member_name:
            # 공백이 아닐때만 추가하도록 함
            if i.text != '':
                member_list.append(i.text)

        print(member_list)
    except Exception as e:
        print(e)

    return member_list


def get_members_info(member_list):
    driver1 = webdriver.Chrome(path, chrome_options=options)
    # driver1 = webdriver.Chrome(path)
    # 엑셀 파일 경로 지정
    file_path = "./memberlist.xlsx"
    write_wb = Workbook()
    write_ws = write_wb.active

    write_ws.cell(1, 1, "닉네임")
    write_ws.cell(1, 2, "무릉도장")
    write_ws.cell(1, 3, "직업")
    write_ws.cell(1, 4, "레벨")

    for i in member_list:
        try:
            url = "https://maple.gg/u/" + i
            driver1.get(url)
            time.sleep(5)
            # 갱신버튼을 누르도록 함
            driver1.find_element_by_xpath('//*[@id="btn-sync"]').click()
            time.sleep(5)
            mureungdojang = driver1.find_element_by_xpath(
                '//*[@id="app"]/div[3]/div/section/div[1]/div[1]/section/div/div[1]/div/h1')
            print('{0} / {1}'.format(i, mureungdojang.text))
            only_mureungdojang = re.sub('[^0-9]', '', str(mureungdojang.text))
            write_ws.cell(member_list.index(i) + 2, 1, i)
            write_ws.cell(member_list.index(i) + 2, 2, only_mureungdojang)
        except Exception as e:
            print(e)

            try:
                old_mureungdojang = driver1.find_element_by_xpath(
                    '//*[@id="app"]/div[3]/div/section/div[1]/div[1]/section/div/div[3]/div/b')
                print('{0} / 예전 무릉 {1}'.format(i, old_mureungdojang.text))
                only_old_mureungdojang = re.sub('[^0-9]', '', str(old_mureungdojang.text))
                write_ws.cell(member_list.index(i) + 2, 1, i)
                write_ws.cell(member_list.index(i) + 2, 2, only_old_mureungdojang)
            except Exception as b:
                print('{0} / 0 층'.format(i))
                write_ws.cell(member_list.index(i) + 2, 1, i)
                write_ws.cell(member_list.index(i) + 2, 2, "0")
                print(b)

        # 직업 부분 추가
        job = driver1.find_element_by_xpath('//*[@id="user-profile"]/section/div/div[2]/div[1]/ul/li[2]')
        print('{0} / {1}'.format(i, job.text))
        job_name = job.text
        if job.text == "다크나이트":
            job_name = "닼나"
        elif job.text == "보우마스터":
            job_name = "보마"
        elif job.text == "패스파인더":
            job_name = "패파"
        elif job.text == "아크메이지(썬,콜)":
            job_name = "썬콜"
        elif job.text == "아크메이지(불,독)":
            job_name = "불독"
        elif job.text == "나이트로드":
            job_name = "나로"
        elif job.text == "듀얼블레이더":
            job_name = "듀블"
        elif job.text == "캐논마스터":
            job_name = "캐슈"
        elif job.text == "와일드헌터":
            job_name = "와헌"
        elif job.text == "배틀메이지":
            job_name = "배메"
        elif job.text == "블래스터":
            job_name = "블래"
        elif job.text == "데몬어벤져":
            job_name = "데벤"
        elif job.text == "데몬슬레이어":
            job_name = "데슬"
        elif job.text == "소울마스터":
            job_name = "소마"
        elif job.text == "윈드브레이커":
            job_name = "윈브"
        elif job.text == "플레임위자드":
            job_name = "플위"
        elif job.text == "나이트워커":
            job_name = "나워"
        elif job.text == "스트라이커":
            job_name = "스커"
        elif job.text == "루미너스":
            job_name = "루미"
        elif job.text == "메르세데스":
            job_name = "메르"
        elif job.text == "엔젤릭버스터":
            job_name = "엔버"
        elif job.text == "키네시스":
            job_name = "키네"

        write_ws.cell(member_list.index(i) + 2, 3, job_name)

        # 레벨 부분 추가
        level = driver1.find_element_by_xpath('//*[@id="user-profile"]/section/div/div[2]/div[1]/ul/li[1]')
        # print('level / {0}'.format(level.text))
        only_level = re.sub('[^0-9]', '', str(level.text))
        print('{0} / {1}'.format(i, only_level))
        write_ws.cell(member_list.index(i) + 2, 4, only_level)

    write_wb.save(file_path)
    driver1.quit()


start = time.time()  # 시작 시간 저장

# 크롬 드라이버 옵션 설정
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
# 혹은 options.add_argument("--disable-gpu")
path = "./chromedriver.exe"
driver = webdriver.Chrome(path, chrome_options=options)
# driver = webdriver.Chrome(path)

# guildname = input("길드명을 입력해주트세요 : ")
guildname = "Adult"
servername = "엘리시움"
memberlist = get_member(guildname, servername)
get_members_info(memberlist)

driver.quit()

print("소요시간 :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간
